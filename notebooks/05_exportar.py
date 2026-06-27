# 05_exportar.py
# Proyecto 3 — Pandemic Education Impact
# Objetivo: consolidar todos los outputs del pipeline en un Excel
# con hojas por etapa + hoja merge-ready para Proyecto 4

import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ── Rutas ──────────────────────────────────────────────────────────────────────
base_dir = os.path.dirname(os.path.abspath(__file__))
out_dir  = os.path.join(base_dir, '..', 'output')

paths = {
    'prep':               os.path.join(out_dir, '01_pandemic_prep.csv'),
    'serie_temporal':     os.path.join(out_dir, '02_serie_temporal.csv'),
    'serie_regional':     os.path.join(out_dir, '02_serie_regional.csv'),
    'evento':             os.path.join(out_dir, '03_evento.csv'),
    'evento_regional':    os.path.join(out_dir, '03_evento_regional.csv'),
    'recuperacion':       os.path.join(out_dir, '04_recuperacion.csv'),
    'recuperacion_reg':   os.path.join(out_dir, '04_recuperacion_regional.csv'),
}

excel_path = os.path.join(out_dir, 'pandemic_education_analisis.xlsx')

# ── Carga ──────────────────────────────────────────────────────────────────────
dfs = {k: pd.read_csv(v) for k, v in paths.items()}
print("CSVs cargados:")
for k, df in dfs.items():
    print(f"  {k}: {df.shape[0]:,} filas × {df.shape[1]} columnas")
print()

# ── Hoja merge-ready para Proyecto 4 ──────────────────────────────────────────
# Columnas clave del índice de recuperación + región, para cruzar con HDI
merge_p4 = dfs['recuperacion'][[
    'country_name', 'region', 'valor_2019',
    'valor_recovery', 'año_recovery',
    'recovery_index', 'delta_vs_2019_pp',
    'estado_recuperacion'
]].copy()

# Agregar baseline y deltas del event study
evento_cols = dfs['evento'][[
    'country_name', 'baseline_mean', 'shock_mean',
    'delta_shock_pp', 'delta_shock_pct',
    'delta_recovery_pp', 'delta_recovery_pct',
    'impacto_shock'
]]
merge_p4 = merge_p4.merge(evento_cols, on='country_name', how='left')

print(f"Hoja merge P4: {merge_p4.shape[0]} países × {merge_p4.shape[1]} columnas")
print(f"Columnas: {merge_p4.columns.tolist()}\n")

# ── Escribir Excel ─────────────────────────────────────────────────────────────
hojas = {
    '01_datos_prep':          dfs['prep'],
    '02_tendencia_global':    dfs['serie_temporal'],
    '02_tendencia_regional':  dfs['serie_regional'],
    '03_event_study_pais':    dfs['evento'],
    '03_event_study_region':  dfs['evento_regional'],
    '04_recovery_pais':       dfs['recuperacion'],
    '04_recovery_region':     dfs['recuperacion_reg'],
    '05_merge_ready_P4':      merge_p4,
}

with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
    for nombre, df in hojas.items():
        df.to_excel(writer, sheet_name=nombre, index=False)
        print(f"  Hoja '{nombre}': {df.shape[0]:,} filas × {df.shape[1]} columnas")

print()

# ── Estilo ─────────────────────────────────────────────────────────────────────
wb = load_workbook(excel_path)

# Paleta
COLOR_HEADER     = '1F3864'  # azul oscuro
COLOR_HEADER_ALT = '2E75B6'  # azul medio (hojas secundarias)
COLOR_ACCENT     = 'D6E4F0'  # azul muy claro (filas pares)
COLOR_MERGE      = '1A5276'  # azul profundo (hoja P4)
FONT_WHITE       = 'FFFFFF'
FONT_DARK        = '1F3864'

def estilo_header(ws, color_hex):
    fill   = PatternFill('solid', fgColor=color_hex)
    font   = Font(bold=True, color=FONT_WHITE, size=10)
    align  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(
        bottom=Side(style='medium', color='FFFFFF')
    )
    for cell in ws[1]:
        cell.fill      = fill
        cell.font      = font
        cell.alignment = align
        cell.border    = border
    ws.row_dimensions[1].height = 30

def estilo_filas(ws, color_alt_hex):
    fill_alt = PatternFill('solid', fgColor=color_alt_hex)
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        for cell in row:
            cell.alignment = Alignment(horizontal='left', vertical='center')
            cell.font      = Font(size=9, color=FONT_DARK)
            if i % 2 == 0:
                cell.fill = fill_alt

def autofit(ws, min_w=10, max_w=40):
    for col in ws.columns:
        max_len = max(
            (len(str(cell.value)) if cell.value is not None else 0)
            for cell in col
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = (
            min(max(max_len + 2, min_w), max_w)
        )

def freeze_header(ws):
    ws.freeze_panes = 'A2'

# Colores por hoja
colores = {
    '01_datos_prep':          COLOR_HEADER,
    '02_tendencia_global':    COLOR_HEADER_ALT,
    '02_tendencia_regional':  COLOR_HEADER_ALT,
    '03_event_study_pais':    COLOR_HEADER,
    '03_event_study_region':  COLOR_HEADER,
    '04_recovery_pais':       COLOR_HEADER,
    '04_recovery_region':     COLOR_HEADER,
    '05_merge_ready_P4':      COLOR_MERGE,
}


for ws in wb.worksheets:
    color = colores.get(ws.title, COLOR_HEADER)
    estilo_header(ws, color)
    estilo_filas(ws, COLOR_ACCENT)
    autofit(ws)
    freeze_header(ws)
    print(f"  Estilo aplicado: {ws.title}")

wb.save(excel_path)
print(f"\nExcel exportado: {excel_path}")

# ── Validación final ───────────────────────────────────────────────────────────
wb2 = load_workbook(excel_path, read_only=True)
print(f"\nHojas en el Excel final: {wb2.sheetnames}")
wb2.close()
